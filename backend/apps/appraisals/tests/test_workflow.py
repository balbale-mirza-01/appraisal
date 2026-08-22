from datetime import timedelta
from io import BytesIO

from django.contrib.auth import get_user_model
from django.core.management import call_command
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.test import APIClient, APITestCase

from apps.appraisals.models import (
    Branch,
    Evaluation,
    EvaluationAssignment,
    EvaluationCycle,
    EvaluationTemplate,
    Region,
    RegionSupervisorAssignment,
)


User = get_user_model()


class EvaluationWorkflowTests(APITestCase):
    @classmethod
    def setUpTestData(cls):
        cls.manager = User.objects.create_user(
            username="manager",
            email="manager@example.com",
            password="StrongPass123!",
            role=User.Role.MARKETING_MANAGER,
        )
        cls.supervisor = User.objects.create_user(
            username="supervisor",
            email="supervisor@example.com",
            password="StrongPass123!",
            role=User.Role.REGION_SUPERVISOR,
        )
        cls.evaluator = User.objects.create_user(
            username="evaluator",
            email="evaluator@example.com",
            password="StrongPass123!",
            role=User.Role.EVALUATOR,
        )
        cls.other_evaluator = User.objects.create_user(
            username="other",
            email="other@example.com",
            password="StrongPass123!",
            role=User.Role.EVALUATOR,
        )
        call_command("seed_initial_template", creator="manager", verbosity=0)
        cls.template = EvaluationTemplate.objects.get(version=1)
        cls.region = Region.objects.create(code="R1", name="منطقه یک")
        cls.other_region = Region.objects.create(code="R2", name="منطقه دو")
        User.objects.filter(pk=cls.evaluator.pk).update(region=cls.region)
        cls.evaluator.refresh_from_db()
        User.objects.filter(pk=cls.other_evaluator.pk).update(region=cls.other_region)
        cls.other_evaluator.refresh_from_db()
        cls.unassigned_evaluator = User.objects.create_user(
            username="newcomer",
            email="newcomer@example.com",
            password="StrongPass123!",
            role=User.Role.EVALUATOR,
            region=cls.region,
        )
        cls.branch = Branch.objects.create(
            region=cls.region, code="B1", name="شعبه یک"
        )
        cls.other_branch = Branch.objects.create(
            region=cls.other_region, code="B2", name="شعبه دو"
        )
        RegionSupervisorAssignment.objects.create(
            region=cls.region, supervisor=cls.supervisor
        )
        today = timezone.localdate()
        cls.cycle = EvaluationCycle.objects.create(
            title="دوره اول",
            template=cls.template,
            start_date=today,
            end_date=today + timedelta(days=30),
            status=EvaluationCycle.Status.ACTIVE,
            created_by=cls.manager,
        )
        cls.assignment = EvaluationAssignment.objects.create(
            cycle=cls.cycle,
            branch=cls.branch,
            evaluator=cls.evaluator,
            assigned_by=cls.supervisor,
            due_date=today + timedelta(days=10),
        )
        cls.other_assignment = EvaluationAssignment.objects.create(
            cycle=cls.cycle,
            branch=cls.other_branch,
            evaluator=cls.other_evaluator,
            assigned_by=cls.manager,
            due_date=today + timedelta(days=10),
        )

    def test_login_returns_access_token_and_refresh_cookie(self):
        self.client.get(reverse("auth-csrf"))
        response = self.client.post(
            reverse("auth-login"),
            {"username": "evaluator", "password": "StrongPass123!"},
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn("access", response.data)
        self.assertIn("appraisal_refresh", response.cookies)
        self.assertNotIn("refresh", response.data)

    def test_login_requires_csrf_when_csrf_checks_are_enabled(self):
        client = APIClient(enforce_csrf_checks=True)
        denied = client.post(
            reverse("auth-login"),
            {"username": "evaluator", "password": "StrongPass123!"},
            format="json",
        )
        self.assertEqual(denied.status_code, status.HTTP_403_FORBIDDEN)

        csrf_response = client.get(reverse("auth-csrf"))
        allowed = client.post(
            reverse("auth-login"),
            {"username": "evaluator", "password": "StrongPass123!"},
            format="json",
            HTTP_X_CSRFTOKEN=csrf_response.data["csrfToken"],
        )
        self.assertEqual(allowed.status_code, status.HTTP_200_OK)

    def test_refresh_rotates_cookie_without_exposing_refresh_token(self):
        self.client.get(reverse("auth-csrf"))
        login_response = self.client.post(
            reverse("auth-login"),
            {"username": "evaluator", "password": "StrongPass123!"},
            format="json",
        )
        original_refresh = login_response.cookies["appraisal_refresh"].value
        refresh_response = self.client.post(reverse("auth-refresh"), {}, format="json")
        self.assertEqual(refresh_response.status_code, status.HTTP_200_OK)
        self.assertIn("access", refresh_response.data)
        self.assertNotIn("refresh", refresh_response.data)
        self.assertIn("appraisal_refresh", refresh_response.cookies)
        self.assertNotEqual(
            original_refresh,
            refresh_response.cookies["appraisal_refresh"].value,
        )

    def test_password_change_revokes_refresh_session(self):
        self.client.get(reverse("auth-csrf"))
        login_response = self.client.post(
            reverse("auth-login"),
            {"username": "evaluator", "password": "StrongPass123!"},
            format="json",
        )
        self.client.credentials(
            HTTP_AUTHORIZATION=f"Bearer {login_response.data['access']}"
        )
        change_response = self.client.post(
            reverse("change-password"),
            {
                "current_password": "StrongPass123!",
                "new_password": "NewStrongPass456!",
            },
            format="json",
        )
        self.assertEqual(change_response.status_code, status.HTTP_200_OK)
        self.client.credentials()
        refresh_response = self.client.post(reverse("auth-refresh"), {}, format="json")
        self.assertEqual(refresh_response.status_code, status.HTTP_401_UNAUTHORIZED)

        new_login = self.client.post(
            reverse("auth-login"),
            {"username": "evaluator", "password": "NewStrongPass456!"},
            format="json",
        )
        self.assertEqual(new_login.status_code, status.HTTP_200_OK)

    def test_dashboards_return_role_specific_sections(self):
        self.client.force_authenticate(self.evaluator)
        evaluator_response = self.client.get(reverse("dashboard"))
        self.assertEqual(evaluator_response.status_code, status.HTTP_200_OK)
        self.assertIn("evaluator_summary", evaluator_response.data)
        self.assertNotIn("supervisor_summary", evaluator_response.data)

        self.client.force_authenticate(self.supervisor)
        supervisor_response = self.client.get(reverse("dashboard"))
        self.assertEqual(supervisor_response.status_code, status.HTTP_200_OK)
        self.assertIn("supervisor_summary", supervisor_response.data)
        self.assertIn("evaluator_workload", supervisor_response.data)

        self.client.force_authenticate(self.manager)
        manager_response = self.client.get(reverse("dashboard"))
        self.assertEqual(manager_response.status_code, status.HTTP_200_OK)
        self.assertIn("manager_summary", manager_response.data)
        self.assertIn("active_cycles", manager_response.data)

    def test_evaluator_only_sees_own_assignment(self):
        self.client.force_authenticate(self.evaluator)
        response = self.client.get(reverse("assignment-list"))
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        ids = [item["id"] for item in response.data["results"]]
        self.assertEqual(ids, [self.assignment.id])

    def test_supervisor_only_sees_assigned_region(self):
        self.client.force_authenticate(self.supervisor)
        response = self.client.get(reverse("assignment-list"))
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        ids = [item["id"] for item in response.data["results"]]
        self.assertEqual(ids, [self.assignment.id])

    def test_complete_evaluation_scores_and_approval_workflow(self):
        answers = [
            {"criterion": criterion.id, "score": 5, "comment": ""}
            for section in self.template.sections.all()
            for criterion in section.criteria.all()
        ]
        self.client.force_authenticate(self.evaluator)
        create_response = self.client.post(
            reverse("evaluation-list"),
            {
                "assignment": self.assignment.id,
                "evaluation_date": timezone.localdate().isoformat(),
                "strengths": "نمونه",
                "answers": answers,
                "opportunities": [],
            },
            format="json",
        )
        self.assertEqual(create_response.status_code, status.HTTP_201_CREATED)
        evaluation_id = create_response.data["id"]
        self.assertEqual(float(create_response.data["total_score"]), 100.0)

        submit_response = self.client.post(
            reverse("evaluation-submit", args=[evaluation_id]), {}, format="json"
        )
        self.assertEqual(submit_response.status_code, status.HTTP_200_OK)
        self.assertEqual(submit_response.data["status"], Evaluation.Status.SUBMITTED)

        self.client.force_authenticate(self.supervisor)
        approve_response = self.client.post(
            reverse("evaluation-approve", args=[evaluation_id]),
            {"comment": "تأیید شد"},
            format="json",
        )
        self.assertEqual(approve_response.status_code, status.HTTP_200_OK)
        self.assertEqual(approve_response.data["status"], Evaluation.Status.APPROVED)

        export_response = self.client.get(
            reverse("evaluation-export-xlsx", args=[evaluation_id])
        )
        self.assertEqual(export_response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            export_response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertTrue(export_response.content.startswith(b"PK"))
        individual_workbook = load_workbook(BytesIO(export_response.content))
        self.assertEqual(
            individual_workbook.sheetnames,
            ["مشخصات", "نتایج بخش‌ها", "پاسخ‌ها", "فرصت‌ها"],
        )

        report_response = self.client.get(reverse("evaluation-report-xlsx"))
        self.assertEqual(report_response.status_code, status.HTTP_200_OK)
        report_workbook = load_workbook(BytesIO(report_response.content))
        self.assertEqual(
            report_workbook.sheetnames,
            ["خلاصه", "مناطق", "شعب", "ارزیابان", "امتیاز بخش‌ها", "جزئیات"],
        )

        self.client.force_authenticate(self.manager)
        reopen_response = self.client.post(
            reverse("evaluation-reopen", args=[evaluation_id]),
            {"comment": "نیاز به اصلاح مدیریتی"},
            format="json",
        )
        self.assertEqual(reopen_response.status_code, status.HTTP_200_OK)
        self.assertEqual(reopen_response.data["status"], Evaluation.Status.RETURNED)

    def test_incomplete_evaluation_cannot_be_submitted(self):
        self.client.force_authenticate(self.evaluator)
        create_response = self.client.post(
            reverse("evaluation-list"),
            {
                "assignment": self.assignment.id,
                "evaluation_date": timezone.localdate().isoformat(),
                "answers": [],
            },
            format="json",
        )
        self.assertEqual(create_response.status_code, status.HTTP_201_CREATED)
        submit_response = self.client.post(
            reverse("evaluation-submit", args=[create_response.data["id"]]),
            {},
            format="json",
        )
        self.assertEqual(submit_response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertIn("answers", submit_response.data)

    def test_supervisor_sees_only_evaluators_from_assigned_region(self):
        """Region supervisor should only see evaluators from their region when listing evaluators."""
        self.client.force_authenticate(self.supervisor)
        response = self.client.get(reverse("evaluator-list"))
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        evaluator_ids = [item["id"] for item in response.data]
        # Supervisor should only see evaluators whose region is in their supervised regions
        self.assertIn(self.evaluator.id, evaluator_ids)
        self.assertNotIn(self.other_evaluator.id, evaluator_ids)

    def test_supervisor_sees_evaluator_with_no_prior_assignments(self):
        """An evaluator with a set region must be visible even before any assignment exists."""
        self.client.force_authenticate(self.supervisor)
        response = self.client.get(reverse("evaluator-list"))
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        evaluator_ids = [item["id"] for item in response.data]
        self.assertIn(self.unassigned_evaluator.id, evaluator_ids)

    def test_manager_sees_all_evaluators(self):
        """Marketing manager should see all evaluators."""
        self.client.force_authenticate(self.manager)
        response = self.client.get(reverse("evaluator-list"))
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        evaluator_ids = [item["id"] for item in response.data]
        # Manager should see all evaluators
        self.assertIn(self.evaluator.id, evaluator_ids)
        self.assertIn(self.other_evaluator.id, evaluator_ids)
