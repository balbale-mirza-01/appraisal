from datetime import timedelta
from io import BytesIO
from urllib.parse import quote

from django.contrib.auth import get_user_model
from django.db.models import Avg, Count, Q
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from rest_framework import mixins, status, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import PermissionDenied, ValidationError
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import (
    AuditEvent,
    Branch,
    Evaluation,
    EvaluationAssignment,
    EvaluationCycle,
    EvaluationTemplate,
    Region,
    RegionSupervisorAssignment,
)
from .permissions import (
    BUSINESS_MANAGERS,
    CanManageAssignments,
    IsMarketingManagerOrAdmin,
    can_review_evaluation,
    supervised_region_ids,
)
from .serializers import (
    AssignmentSerializer,
    BranchSerializer,
    CycleSerializer,
    EvaluationSerializer,
    RegionSerializer,
    RegionSupervisorAssignmentSerializer,
    ReviewSerializer,
    TemplateSerializer,
    UserSerializer,
)
from .services import (
    aggregate_status_counts,
    recalculate_evaluation,
    reopen_evaluation,
    review_evaluation,
    submit_evaluation,
)


User = get_user_model()


def visible_assignments(user):
    queryset = EvaluationAssignment.objects.select_related(
        "cycle",
        "cycle__template",
        "branch",
        "branch__region",
        "evaluator",
        "assigned_by",
    )
    if user.role in BUSINESS_MANAGERS:
        return queryset
    if user.role == User.Role.REGION_SUPERVISOR:
        return queryset.filter(branch__region_id__in=supervised_region_ids(user))
    return queryset.filter(evaluator=user)


def visible_evaluations(user):
    queryset = Evaluation.objects.select_related(
        "assignment",
        "assignment__cycle",
        "assignment__cycle__template",
        "assignment__branch",
        "assignment__branch__region",
        "assignment__evaluator",
        "assignment__assigned_by",
        "reviewed_by",
    ).prefetch_related("answers", "opportunities")
    if user.role in BUSINESS_MANAGERS:
        return queryset
    if user.role == User.Role.REGION_SUPERVISOR:
        return queryset.filter(
            assignment__branch__region_id__in=supervised_region_ids(user)
        )
    return queryset.filter(assignment__evaluator=user)


class RegionViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = RegionSerializer
    pagination_class = None

    def get_queryset(self):
        queryset = Region.objects.filter(is_active=True)
        user = self.request.user
        if user.role == User.Role.REGION_SUPERVISOR:
            queryset = queryset.filter(id__in=supervised_region_ids(user))
        elif user.role == User.Role.EVALUATOR:
            queryset = queryset.filter(
                branches__evaluation_assignments__evaluator=user
            ).distinct()
        return queryset


class BranchViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = BranchSerializer
    pagination_class = None
    filterset_fields = ("region",)
    search_fields = ("name", "code", "manager_name")
    ordering_fields = ("name", "code")

    def get_queryset(self):
        queryset = Branch.objects.filter(is_active=True).select_related("region")
        user = self.request.user
        if user.role == User.Role.REGION_SUPERVISOR:
            queryset = queryset.filter(region_id__in=supervised_region_ids(user))
        elif user.role == User.Role.EVALUATOR:
            queryset = queryset.filter(evaluation_assignments__evaluator=user).distinct()
        return queryset


class TemplateViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = TemplateSerializer
    pagination_class = None

    def get_queryset(self):
        return EvaluationTemplate.objects.filter(
            status=EvaluationTemplate.Status.PUBLISHED
        ).prefetch_related("sections__criteria")

    @action(detail=False, methods=("get",))
    def current(self, request):
        template = self.get_queryset().order_by("-version").first()
        if not template:
            return Response(
                {"detail": "الگوی منتشرشده‌ای وجود ندارد."},
                status=status.HTTP_404_NOT_FOUND,
            )
        return Response(self.get_serializer(template).data)


class CycleViewSet(viewsets.ModelViewSet):
    serializer_class = CycleSerializer
    pagination_class = None
    filterset_fields = ("status", "template")
    search_fields = ("title",)
    ordering_fields = ("start_date", "end_date", "title")

    def get_queryset(self):
        queryset = EvaluationCycle.objects.select_related("template", "created_by")
        user = self.request.user
        if user.role in BUSINESS_MANAGERS:
            return queryset
        if user.role == User.Role.REGION_SUPERVISOR:
            return queryset.filter(
                Q(status=EvaluationCycle.Status.ACTIVE)
                | Q(
                    assignments__branch__region_id__in=supervised_region_ids(user)
                )
            ).distinct()
        return queryset.filter(assignments__evaluator=user).distinct()

    def get_permissions(self):
        if self.action in {"create", "update", "partial_update", "destroy"}:
            return [IsMarketingManagerOrAdmin()]
        return [IsAuthenticated()]


class EvaluatorViewSet(mixins.ListModelMixin, viewsets.GenericViewSet):
    serializer_class = UserSerializer
    permission_classes = (CanManageAssignments,)
    pagination_class = None
    search_fields = ("username", "first_name", "last_name", "employee_number")

    def get_queryset(self):
        queryset = User.objects.filter(
            role=User.Role.EVALUATOR, is_active=True
        ).order_by("last_name", "first_name", "username")
        user = self.request.user
        if user.role == User.Role.REGION_SUPERVISOR:
            queryset = queryset.filter(
                evaluation_assignments__branch__region_id__in=supervised_region_ids(
                    user
                )
            ).distinct()
        return queryset


class AssignmentViewSet(viewsets.ModelViewSet):
    serializer_class = AssignmentSerializer
    filterset_fields = ("cycle", "branch", "evaluator", "status")
    search_fields = (
        "branch__name",
        "branch__code",
        "evaluator__first_name",
        "evaluator__last_name",
        "evaluator__employee_number",
    )
    ordering_fields = ("due_date", "created_at", "status")

    def get_queryset(self):
        return visible_assignments(self.request.user)

    def get_permissions(self):
        if self.action in {"create", "update", "partial_update", "destroy"}:
            return [CanManageAssignments()]
        return [IsAuthenticated()]

    def perform_destroy(self, instance):
        if hasattr(instance, "evaluation"):
            raise ValidationError("تخصیصی که ارزیابی آن آغاز شده قابل حذف نیست.")
        instance.delete()


class EvaluationViewSet(viewsets.ModelViewSet):
    serializer_class = EvaluationSerializer
    filterset_fields = (
        "status",
        "assignment__cycle",
        "assignment__branch",
        "assignment__branch__region",
        "assignment__evaluator",
    )
    search_fields = (
        "assignment__branch__name",
        "assignment__branch__code",
        "assignment__evaluator__first_name",
        "assignment__evaluator__last_name",
    )
    ordering_fields = (
        "evaluation_date",
        "created_at",
        "updated_at",
        "total_score",
        "status",
    )

    def get_queryset(self):
        return visible_evaluations(self.request.user)

    def perform_create(self, serializer):
        evaluation = serializer.save()
        AuditEvent.objects.create(
            evaluation=evaluation,
            actor=self.request.user,
            action="created",
            metadata={"assignment_id": evaluation.assignment_id},
        )

    def _ensure_editable(self, evaluation):
        if evaluation.assignment.evaluator_id != self.request.user.id:
            raise PermissionDenied("فقط ارزیاب تخصیص‌یافته می‌تواند فرم را ویرایش کند.")
        if evaluation.status not in (
            Evaluation.Status.DRAFT,
            Evaluation.Status.RETURNED,
        ):
            raise ValidationError("ارزیابی در وضعیت فعلی قابل ویرایش نیست.")

    def update(self, request, *args, **kwargs):
        self._ensure_editable(self.get_object())
        return super().update(request, *args, **kwargs)

    def partial_update(self, request, *args, **kwargs):
        self._ensure_editable(self.get_object())
        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        evaluation = self.get_object()
        self._ensure_editable(evaluation)
        assignment = evaluation.assignment
        response = super().destroy(request, *args, **kwargs)
        assignment.status = EvaluationAssignment.Status.ASSIGNED
        assignment.save(update_fields=["status", "updated_at"])
        return response

    @action(detail=True, methods=("post",))
    def submit(self, request, pk=None):
        evaluation = submit_evaluation(self.get_object(), request.user)
        return Response(self.get_serializer(evaluation).data)

    @action(detail=True, methods=("post",))
    def approve(self, request, pk=None):
        evaluation = self.get_object()
        if not can_review_evaluation(request.user, evaluation):
            raise PermissionDenied("مجوز بررسی این ارزیابی را ندارید.")
        serializer = ReviewSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        review_evaluation(
            evaluation,
            request.user,
            approve=True,
            comment=serializer.validated_data.get("comment", ""),
        )
        return Response(self.get_serializer(evaluation).data)

    @action(detail=True, methods=("post",))
    def return_for_correction(self, request, pk=None):
        evaluation = self.get_object()
        if not can_review_evaluation(request.user, evaluation):
            raise PermissionDenied("مجوز بررسی این ارزیابی را ندارید.")
        serializer = ReviewSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        comment = serializer.validated_data.get("comment", "").strip()
        if not comment:
            raise ValidationError({"comment": "برای بازگشت ارزیابی درج توضیح الزامی است."})
        review_evaluation(
            evaluation, request.user, approve=False, comment=comment
        )
        return Response(self.get_serializer(evaluation).data)

    @action(detail=True, methods=("post",))
    def recalculate(self, request, pk=None):
        evaluation = self.get_object()
        if request.user.role not in BUSINESS_MANAGERS:
            raise PermissionDenied("فقط مدیران مجاز به محاسبه مجدد هستند.")
        recalculate_evaluation(evaluation)
        return Response(self.get_serializer(evaluation).data)

    @action(detail=True, methods=("post",))
    def reopen(self, request, pk=None):
        evaluation = self.get_object()
        if request.user.role not in BUSINESS_MANAGERS:
            raise PermissionDenied("فقط مدیر بازاریابی یا مدیر سامانه مجاز است.")
        serializer = ReviewSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        evaluation = reopen_evaluation(
            evaluation,
            request.user,
            serializer.validated_data.get("comment", ""),
        )
        return Response(self.get_serializer(evaluation).data)

    @action(detail=True, methods=("get",), url_path="export-xlsx")
    def export_xlsx(self, request, pk=None):
        evaluation = self.get_object()
        AuditEvent.objects.create(
            evaluation=evaluation,
            actor=request.user,
            action="evaluation_exported_xlsx",
            metadata={},
        )
        return individual_evaluation_workbook(evaluation)


class RegionSupervisorAssignmentViewSet(viewsets.ModelViewSet):
    serializer_class = RegionSupervisorAssignmentSerializer
    permission_classes = (IsMarketingManagerOrAdmin,)
    queryset = RegionSupervisorAssignment.objects.select_related(
        "region", "supervisor"
    )
    filterset_fields = ("region", "supervisor", "is_active")


class DashboardView(APIView):
    def get(self, request):
        assignments = visible_assignments(request.user)
        evaluations = visible_evaluations(request.user)
        today = timezone.localdate()
        payload = {
            "role": request.user.role,
            "assignment_counts": aggregate_status_counts(assignments),
            "evaluation_counts": aggregate_status_counts(evaluations),
            "total_assignments": assignments.count(),
            "total_evaluations": evaluations.count(),
            "average_score": evaluations.filter(
                status=Evaluation.Status.APPROVED
            ).aggregate(value=Avg("total_score"))["value"],
            "due_soon": assignments.filter(
                due_date__gte=today,
                due_date__lte=today + timedelta(days=7),
            )
            .exclude(status=EvaluationAssignment.Status.APPROVED)
            .count(),
        }

        if request.user.role in {
            User.Role.REGION_SUPERVISOR,
            User.Role.MARKETING_MANAGER,
            User.Role.ADMIN,
        }:
            regional = (
                evaluations.values(
                    "assignment__branch__region_id",
                    "assignment__branch__region__name",
                )
                .annotate(
                    evaluation_count=Count("id"),
                    approved_count=Count(
                        "id", filter=Q(status=Evaluation.Status.APPROVED)
                    ),
                    average_score=Avg(
                        "total_score", filter=Q(status=Evaluation.Status.APPROVED)
                    ),
                )
                .order_by("assignment__branch__region__name")
            )
            payload["regions"] = list(regional)

        if request.user.role == User.Role.EVALUATOR:
            approved_count = assignments.filter(
                status=EvaluationAssignment.Status.APPROVED
            ).count()
            payload["evaluator_summary"] = {
                "approved": approved_count,
                "returned": assignments.filter(
                    status=EvaluationAssignment.Status.RETURNED
                ).count(),
                "overdue": assignments.filter(due_date__lt=today)
                .exclude(status=EvaluationAssignment.Status.APPROVED)
                .count(),
                "completion_rate": round(
                    approved_count / assignments.count() * 100, 1
                )
                if assignments.exists()
                else 0,
            }
            payload["action_required"] = AssignmentSerializer(
                assignments.filter(
                    status__in=[
                        EvaluationAssignment.Status.ASSIGNED,
                        EvaluationAssignment.Status.IN_PROGRESS,
                        EvaluationAssignment.Status.RETURNED,
                    ]
                )[:10],
                many=True,
            ).data
        else:
            payload["waiting_for_review"] = EvaluationSerializer(
                evaluations.filter(status=Evaluation.Status.SUBMITTED)[:10],
                many=True,
                context={"request": request},
            ).data

        if request.user.role == User.Role.REGION_SUPERVISOR:
            region_ids = supervised_region_ids(request.user)
            payload["supervisor_summary"] = {
                "regions": Region.objects.filter(id__in=region_ids).count(),
                "branches": Branch.objects.filter(
                    region_id__in=region_ids, is_active=True
                ).count(),
                "active_evaluators": assignments.values("evaluator_id")
                .distinct()
                .count(),
                "overdue": assignments.filter(due_date__lt=today)
                .exclude(status=EvaluationAssignment.Status.APPROVED)
                .count(),
            }
            payload["evaluator_workload"] = list(
                assignments.values(
                    "evaluator_id",
                    "evaluator__username",
                    "evaluator__first_name",
                    "evaluator__last_name",
                )
                .annotate(
                    total=Count("id"),
                    in_progress=Count(
                        "id",
                        filter=Q(
                            status__in=[
                                EvaluationAssignment.Status.ASSIGNED,
                                EvaluationAssignment.Status.IN_PROGRESS,
                                EvaluationAssignment.Status.RETURNED,
                            ]
                        ),
                    ),
                    waiting_review=Count(
                        "id",
                        filter=Q(status=EvaluationAssignment.Status.SUBMITTED),
                    ),
                    approved=Count(
                        "id",
                        filter=Q(status=EvaluationAssignment.Status.APPROVED),
                    ),
                    overdue=Count(
                        "id",
                        filter=Q(due_date__lt=today)
                        & ~Q(status=EvaluationAssignment.Status.APPROVED),
                    ),
                )
                .order_by("evaluator__last_name", "evaluator__first_name")
            )

        if request.user.role == User.Role.MARKETING_MANAGER:
            payload["manager_summary"] = {
                "regions": Region.objects.filter(is_active=True).count(),
                "branches": Branch.objects.filter(is_active=True).count(),
                "evaluators": User.objects.filter(
                    role=User.Role.EVALUATOR, is_active=True
                ).count(),
                "active_cycles": EvaluationCycle.objects.filter(
                    status=EvaluationCycle.Status.ACTIVE
                ).count(),
            }
            payload["active_cycles"] = list(
                EvaluationCycle.objects.filter(status=EvaluationCycle.Status.ACTIVE)
                .annotate(
                    assignment_count=Count("assignments", distinct=True),
                    submitted_count=Count(
                        "assignments__evaluation",
                        filter=Q(
                            assignments__evaluation__status=Evaluation.Status.SUBMITTED
                        ),
                        distinct=True,
                    ),
                    approved_count=Count(
                        "assignments__evaluation",
                        filter=Q(
                            assignments__evaluation__status=Evaluation.Status.APPROVED
                        ),
                        distinct=True,
                    ),
                    average_score=Avg(
                        "assignments__evaluation__total_score",
                        filter=Q(
                            assignments__evaluation__status=Evaluation.Status.APPROVED
                        ),
                    ),
                )
                .values(
                    "id",
                    "title",
                    "start_date",
                    "end_date",
                    "assignment_count",
                    "submitted_count",
                    "approved_count",
                    "average_score",
                )
                .order_by("-start_date")
            )

        if request.user.role == User.Role.ADMIN:
            payload["admin_summary"] = {
                "active_users": User.objects.filter(is_active=True).count(),
                "evaluators": User.objects.filter(
                    role=User.Role.EVALUATOR, is_active=True
                ).count(),
                "supervisors": User.objects.filter(
                    role=User.Role.REGION_SUPERVISOR, is_active=True
                ).count(),
                "regions": Region.objects.count(),
                "branches": Branch.objects.count(),
                "templates": EvaluationTemplate.objects.count(),
            }
        return Response(payload)


class EvaluationReportExportView(APIView):
    def get(self, request):
        queryset = visible_evaluations(request.user)
        for field, lookup in {
            "cycle": "assignment__cycle_id",
            "region": "assignment__branch__region_id",
            "branch": "assignment__branch_id",
            "status": "status",
        }.items():
            value = request.query_params.get(field)
            if value:
                queryset = queryset.filter(**{lookup: value})
        AuditEvent.objects.create(
            actor=request.user,
            action="report_exported_xlsx",
            metadata={
                "filters": {
                    key: request.query_params.get(key)
                    for key in ("cycle", "region", "branch", "status")
                    if request.query_params.get(key)
                },
                "evaluation_count": queryset.count(),
            },
        )
        return report_workbook(queryset)


HEADER_FILL = PatternFill("solid", fgColor="662D91")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def style_sheet(sheet):
    sheet.sheet_view.rightToLeft = True
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = "A2"
    for column in sheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 55)
        sheet.column_dimensions[column[0].column_letter].width = width


def workbook_response(workbook, filename):
    output = BytesIO()
    workbook.save(output)
    response = HttpResponse(
        output.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = (
        f"attachment; filename=appraisal.xlsx; filename*=UTF-8''{quote(filename)}"
    )
    return response


def individual_evaluation_workbook(evaluation):
    workbook = Workbook()
    summary = workbook.active
    summary.title = "مشخصات"
    summary.append(["عنوان", "مقدار"])
    rows = [
        ("شعبه", evaluation.branch.name),
        ("کد شعبه", evaluation.branch.code),
        ("منطقه", evaluation.branch.region.name),
        ("ارزیاب", str(evaluation.evaluator)),
        ("دوره", evaluation.assignment.cycle.title),
        ("تاریخ ارزیابی", evaluation.evaluation_date.isoformat()),
        ("وضعیت", evaluation.get_status_display()),
        ("امتیاز کل", float(evaluation.total_score)),
        ("نقاط قوت", evaluation.strengths),
        ("نقاط قابل بهبود", evaluation.improvements),
        ("فرصت‌های بازار", evaluation.market_opportunities),
        ("نیازهای شعبه", evaluation.branch_needs),
        ("نظر بازبین", evaluation.review_comment),
    ]
    for row in rows:
        summary.append(row)
    style_sheet(summary)

    section_sheet = workbook.create_sheet("نتایج بخش‌ها")
    section_sheet.append(["بخش", "امتیاز", "وزن بخش", "درصد تحقق"])
    for section_model in evaluation.assignment.cycle.template.sections.all():
        section = evaluation.section_scores.get(str(section_model.id))
        if not section:
            continue
        section_sheet.append(
            [
                section["title"],
                section["score"],
                section["weight"],
                section["percentage"],
            ]
        )
    style_sheet(section_sheet)

    answers_sheet = workbook.create_sheet("پاسخ‌ها")
    answers_sheet.append(
        ["بخش", "معیار", "امتیاز", "وزن معیار", "امتیاز وزنی", "توضیح"]
    )
    for answer in evaluation.answers.select_related(
        "criterion", "criterion__section"
    ).order_by("criterion__section__order", "criterion__order"):
        answers_sheet.append(
            [
                answer.criterion.section.title,
                answer.criterion.text,
                answer.score,
                float(answer.criterion.weight),
                float(answer.weighted_score),
                answer.comment,
            ]
        )
    style_sheet(answers_sheet)

    opportunity_sheet = workbook.create_sheet("فرصت‌ها")
    opportunity_sheet.append(
        [
            "نام سازمان",
            "تعداد کارکنان",
            "انواع فرصت",
            "مسئول پیگیری",
            "وضعیت",
            "تاریخ هدف",
            "توضیحات",
        ]
    )
    for opportunity in evaluation.opportunities.all():
        opportunity_sheet.append(
            [
                opportunity.organization_name,
                opportunity.employee_count,
                "، ".join(opportunity.opportunity_types),
                opportunity.responsible_person,
                opportunity.status,
                opportunity.target_date.isoformat()
                if opportunity.target_date
                else "",
                opportunity.notes,
            ]
        )
    style_sheet(opportunity_sheet)
    return workbook_response(
        workbook,
        f"ارزیابی_{evaluation.branch.code}_{evaluation.evaluation_date}.xlsx",
    )


def report_workbook(queryset):
    evaluations = list(
        queryset.select_related(
            "assignment__branch__region",
            "assignment__cycle",
            "assignment__evaluator",
        ).order_by("assignment__branch__region__name", "assignment__branch__code")
    )
    workbook = Workbook()
    summary = workbook.active
    summary.title = "خلاصه"
    approved = [
        evaluation
        for evaluation in evaluations
        if evaluation.status == Evaluation.Status.APPROVED
    ]
    average_score = (
        sum(float(evaluation.total_score) for evaluation in approved) / len(approved)
        if approved
        else 0
    )
    status_counts = {
        status_value: sum(
            evaluation.status == status_value for evaluation in evaluations
        )
        for status_value, _ in Evaluation.Status.choices
    }
    summary.append(["شاخص", "مقدار"])
    summary_rows = [
        ("تعداد کل ارزیابی‌ها", len(evaluations)),
        ("پیش‌نویس", status_counts[Evaluation.Status.DRAFT]),
        ("در انتظار بررسی", status_counts[Evaluation.Status.SUBMITTED]),
        ("برگشت داده شده", status_counts[Evaluation.Status.RETURNED]),
        ("تأیید شده", status_counts[Evaluation.Status.APPROVED]),
        ("میانگین امتیاز تأییدشده", round(average_score, 2)),
        ("زمان تهیه گزارش", timezone.localtime().isoformat()),
    ]
    for row in summary_rows:
        summary.append(row)
    style_sheet(summary)

    region_sheet = workbook.create_sheet("مناطق")
    region_sheet.append(["منطقه", "کل ارزیابی", "تأییدشده", "میانگین امتیاز"])
    for label, stats in aggregate_workbook_group(
        evaluations,
        lambda evaluation: evaluation.branch.region.name,
    ):
        region_sheet.append([label, stats["total"], stats["approved"], stats["average"]])
    style_sheet(region_sheet)

    branch_sheet = workbook.create_sheet("شعب")
    branch_sheet.append(
        ["کد شعبه", "شعبه", "منطقه", "کل ارزیابی", "تأییدشده", "میانگین امتیاز"]
    )
    for _, stats in aggregate_workbook_group(
        evaluations,
        lambda evaluation: (
            evaluation.branch.code,
            evaluation.branch.name,
            evaluation.branch.region.name,
        ),
    ):
        code, name, region_name = stats["key"]
        branch_sheet.append(
            [
                code,
                name,
                region_name,
                stats["total"],
                stats["approved"],
                stats["average"],
            ]
        )
    style_sheet(branch_sheet)

    evaluator_sheet = workbook.create_sheet("ارزیابان")
    evaluator_sheet.append(
        ["ارزیاب", "شماره پرسنلی", "کل ارزیابی", "تأییدشده", "میانگین امتیاز"]
    )
    for _, stats in aggregate_workbook_group(
        evaluations,
        lambda evaluation: (
            str(evaluation.evaluator),
            evaluation.evaluator.employee_number or "",
        ),
    ):
        evaluator_name, employee_number = stats["key"]
        evaluator_sheet.append(
            [
                evaluator_name,
                employee_number,
                stats["total"],
                stats["approved"],
                stats["average"],
            ]
        )
    style_sheet(evaluator_sheet)

    section_sheet = workbook.create_sheet("امتیاز بخش‌ها")
    section_sheet.append(
        [
            "منطقه",
            "کد شعبه",
            "شعبه",
            "دوره",
            "بخش",
            "امتیاز",
            "وزن",
            "درصد تحقق",
        ]
    )
    for evaluation in evaluations:
        for section in evaluation.section_scores.values():
            section_sheet.append(
                [
                    evaluation.branch.region.name,
                    evaluation.branch.code,
                    evaluation.branch.name,
                    evaluation.assignment.cycle.title,
                    section["title"],
                    section["score"],
                    section["weight"],
                    section["percentage"],
                ]
            )
    style_sheet(section_sheet)

    detail_sheet = workbook.create_sheet("جزئیات")
    detail_sheet.append(
        [
            "منطقه",
            "کد شعبه",
            "شعبه",
            "دوره",
            "ارزیاب",
            "تاریخ",
            "وضعیت",
            "امتیاز کل",
            "زمان ارسال",
            "زمان بررسی",
        ]
    )
    for evaluation in evaluations:
        detail_sheet.append(
            [
                evaluation.branch.region.name,
                evaluation.branch.code,
                evaluation.branch.name,
                evaluation.assignment.cycle.title,
                str(evaluation.evaluator),
                evaluation.evaluation_date.isoformat(),
                evaluation.get_status_display(),
                float(evaluation.total_score),
                evaluation.submitted_at.isoformat() if evaluation.submitted_at else "",
                evaluation.reviewed_at.isoformat() if evaluation.reviewed_at else "",
            ]
        )
    style_sheet(detail_sheet)
    return workbook_response(workbook, "گزارش_ارزیابی_شعب.xlsx")


def aggregate_workbook_group(evaluations, key_function):
    groups = {}
    for evaluation in evaluations:
        key = key_function(evaluation)
        stats = groups.setdefault(
            key,
            {"key": key, "total": 0, "approved": 0, "scores": []},
        )
        stats["total"] += 1
        if evaluation.status == Evaluation.Status.APPROVED:
            stats["approved"] += 1
            stats["scores"].append(float(evaluation.total_score))

    rows = []
    for key, stats in groups.items():
        stats["average"] = (
            round(sum(stats["scores"]) / len(stats["scores"]), 2)
            if stats["scores"]
            else 0
        )
        label = key if isinstance(key, str) else " - ".join(str(item) for item in key)
        rows.append((label, stats))
    return sorted(rows, key=lambda row: row[0])
